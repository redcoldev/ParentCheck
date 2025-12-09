import os
import json
import csv
from datetime import datetime
from functools import wraps

from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, session
)
from flask_login import (
    LoginManager, UserMixin, login_user,
    logout_user, login_required, current_user
)
from openpyxl import load_workbook
import psycopg
import requests


SANCTION_DATASETS = {
    "us_ofac_sdn",
    "us_ofac_cons",
    "eu_fsf",
    "uk_hmt_sanctions",
    "un_sc_sanctions",
    "au_dfat_sanctions",
    "ca_sema_sanctions",
    "ua_sanctions",
    "ru_ns_sanctions"
}

# =====================================================================
# CONFIG
# =====================================================================

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "changeme")

DB_URL = os.environ.get("DATABASE_URL")
OPEN_SANCTIONS_KEY = os.environ.get("OPEN_SANCTIONS_KEY")

if not DB_URL:
    raise RuntimeError("DATABASE_URL missing")

DB_URL = DB_URL.replace("postgres://", "postgresql://")


# =====================================================================
# LOGIN
# =====================================================================

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


class User(UserMixin):
    def __init__(self, id_, email, school_name):
        self.id = id_
        self.email = email
        self.school_name = school_name


@login_manager.user_loader
def load_user(user_id):
    conn = psycopg.connect(DB_URL)
    cur = conn.cursor()
    cur.execute("SELECT id, email, school_name FROM users WHERE id=%s", (user_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    return User(*row) if row else None


# =====================================================================
# DB HELPERS
# =====================================================================

def get_db():
    conn = psycopg.connect(DB_URL)
    return conn, conn.cursor()


# =====================================================================
# FILE LOADER
# =====================================================================

def load_uploaded_file(file):
    name = file.filename.lower()

    if name.endswith(".csv"):
        data = file.read().decode("utf-8", errors="ignore").splitlines()
        return list(csv.reader(data))

    if name.endswith(".xlsx"):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        return [[str(c or "").strip() for c in row] for row in ws.iter_rows(values_only=True)]

    raise ValueError("Invalid file")


# =====================================================================
# NORMALISE ROWS
# =====================================================================

def normalise_rows(rows):
    if not rows:
        return []

    # header detection
    if any(ch.isalpha() for ch in "".join(rows[0])):
        rows = rows[1:]

    # drop empty rows
    rows = [r for r in rows if any(cell.strip() for cell in r)]

    clean = []
    for r in rows:
        while len(r) < 4:
            r.append("")
        clean.append({
            "first_name": r[0].strip(),
            "last_name": r[1].strip(),
            "country_of_citizenship": r[2].strip(),
            "dob": r[3].strip(),
        })
    return clean


# =====================================================================
# ROUTES
# =====================================================================

@app.route("/")
def index():
    if not current_user.is_authenticated:
        return redirect("/login")
    return redirect("/dashboard")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()

        conn, cur = get_db()
        cur.execute("""
            SELECT id, email, password, school_name
            FROM users WHERE email=%s
        """, (email,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if not user or password != user[2]:
            flash("Invalid credentials", "danger")
            return redirect("/login")

        uid, uemail, _, school = user
        login_user(User(uid, uemail, school))
        return redirect("/dashboard")

    return render_template("login.html")


@app.route("/dashboard")
@login_required
def dashboard():
    conn, cur = get_db()
    cur.execute("""
        SELECT id, filename, uploaded_at
        FROM batches
        WHERE user_id=%s
        ORDER BY uploaded_at DESC
    """, (current_user.id,))
    batches = cur.fetchall()
    cur.close()
    conn.close()

    return render_template("dashboard.html", batches=batches)


@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Upload a file")
            return render_template("upload.html")

        try:
            rows_raw = load_uploaded_file(file)
        except:
            flash("Invalid file")
            return render_template("upload.html")

        rows_clean = normalise_rows(rows_raw)

        conn, cur = get_db()
        cur.execute("""
            INSERT INTO batches (user_id, filename, preview_data, total_rows)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (
            current_user.id,
            file.filename,
            json.dumps(rows_clean[:10]),
            len(rows_clean)
        ))
        batch_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()

        session[f"batch_{batch_id}_rows"] = rows_clean
        return redirect(f"/preview/{batch_id}")

    return render_template("upload.html")


@app.route("/preview/<int:batch_id>")
@login_required
def preview(batch_id):
    conn, cur = get_db()
    cur.execute("""
        SELECT filename, preview_data, total_rows
        FROM batches WHERE id=%s
    """, (batch_id,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        flash("Batch not found")
        return redirect("/dashboard")

    filename, preview_data, total = row

    return render_template(
        "preview.html",
        filename=filename,
        preview=preview_data,
        total=total,
        batch_id=batch_id
    )


@app.route("/processing/<int:batch_id>")
@login_required
def processing(batch_id):
    rows = session.get(f"batch_{batch_id}_rows", [])
    return render_template(
        "processing.html",
        rows_json=json.dumps(rows),
        batch_id=batch_id
    )


# =====================================================================
# API SCREEN — FULL ENTITY PROFILE FOR EXPANDED VIEW
# =====================================================================

@app.route("/api/screen", methods=["POST"])
def api_screen():
    data = request.json

    first = data.get("first_name", "").strip()
    last = data.get("last_name", "").strip()
    raw_dob = data.get("dob", "").strip()

    def normalise_dob(d):
        if not d:
            return None
        d = d.replace("/", "-")
        parts = d.split("-")
        if len(parts) == 3:
            dd, mm, yyyy = parts
            return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"
        return None

    dob = normalise_dob(raw_dob)

    headers = {
        "Authorization": f"ApiKey {OPEN_SANCTIONS_KEY}",
        "Content-Type": "application/json",
    }

    payload = {
        "queries": {
            "q": {
                "schema": "Person",
                "properties": {
                    "firstName": [first],
                    "lastName": [last]
                }
            }
        }
    }

    if dob:
        payload["queries"]["q"]["properties"]["birthDate"] = [dob]

    try:
        resp = requests.post(
            "https://api.opensanctions.org/match/default",
            headers=headers,
            json=payload,
            timeout=12
        )
        os_json = resp.json()
    except Exception as e:
        return {"risk": "Error", "summary": str(e)}

    results = os_json.get("responses", {}).get("q", {}).get("results", [])

    # Matching helpers
    def names_match(props):
        aliases = props.get("alias", []) + props.get("name", [])
        for a in aliases:
            parts = a.lower().split()
            if len(parts) >= 2 and parts[0] == first.lower() and parts[-1] == last.lower():
                return True
        return False

    def dob_matches(u_dob, birth_dates):
        if not u_dob:
            return True
        if not birth_dates:
            return True
        yyyy = u_dob[:4]
        for bd in birth_dates:
            if not bd:
                continue
            if len(bd) >= 10 and bd == u_dob:
                return True
            if bd[:4] == yyyy:
                return True
        return False

    # Evaluate matches
    for m in results:
        props = m.get("properties", {})
        score = m.get("score", 0)
        datasets = m.get("datasets", [])

        if score != 1.0:
            continue
        if not any(ds in SANCTION_DATASETS for ds in datasets):
            continue
        if not names_match(props):
            continue
        if not dob_matches(dob, props.get("birthDate", [])):
            continue

        # Clean sanctions entries
        sanctions = []
        for s in props.get("sanctions", []):
            sanctions.append({
                "program": s.get("program"),
                "authority": s.get("authority"),
                "listingDate": s.get("listingDate"),
                "reason": (s.get("reason") or "Reason not provided")[:200]
            })

        # Full profile — send ALL fields
        return {
            "risk": "Match",
            "summary": f"{first} {last} appears on sanctions lists",
            "datasets": datasets,
            "short_profile": props.get("summary", "")[:200],

            # sanctions table (frontend uses it)
            "sanctions": sanctions,

            # FULL PROPERTIES for dynamic OS-style sections
            "props": props
        }


    return {"risk": "Clear", "summary": "No sanctions match."}


# =====================================================================
# MAIN BATCH PROCESSOR (unchanged from prior version)
# =====================================================================

def process_batch(batch_id, rows):
    API_KEY = os.environ.get("OPEN_SANCTIONS_KEY")
    headers = {"Authorization": f"ApiKey {API_KEY}", "Content-Type": "application/json"}

    def dob_matches(user_dob, os_birth_dates):
        digits = ''.join(ch for ch in user_dob if ch.isdigit())
        if len(digits) != 8:
            return False
        dd, mm, yyyy = digits[:2], digits[2:4], digits[4:]
        for bd in os_birth_dates:
            if not bd:
                continue
            if bd[:4] != yyyy:
                continue
            if len(bd) < 10:
                return True
            if bd[8:10] == dd and bd[5:7] == mm:
                return True
        return False

    def citizenship_matches(country, props):
        if not country:
            return True
        nat = props.get("nationality", []) + props.get("citizenship", [])
        nat = [x.lower() for x in nat]
        return country.lower() in nat

    batch_results = []

    for idx, r in enumerate(rows):
        query_id = f"row{idx}"
        properties = {"firstName": [r["first_name"]], "lastName": [r["last_name"]]}

        if r.get("dob"):
            properties["birthDate"] = [r["dob"]]
        if r.get("country_of_citizenship"):
            properties["country"] = [r["country_of_citizenship"]]

        payload = {"queries": {query_id: {"schema": "Person", "properties": properties}}}

        try:
            resp = requests.post(
                "https://api.opensanctions.org/match/sanctions",
                headers=headers,
                json=payload,
                timeout=12
            )
            os_json = resp.json()
        except Exception:
            os_json = {"error": "Failed OS request"}

        results_raw = os_json.get("responses", {}).get(query_id, {}).get("results", [])
        true_matches = []

        for m in results_raw:
            score = m.get("score", 0)
            props = m.get("properties", {})
            if score < 0.75:
                continue
            if not citizenship_matches(r.get("country_of_citizenship"), props):
                continue
            if not dob_matches(r.get("dob"), props.get("birthDate", [])):
                continue
            true_matches.append(m)

        risk = "High" if true_matches else "Clear"

        batch_results.append({
            "batch_id": batch_id,
            "first_name": r["first_name"],
            "last_name": r["last_name"],
            "dob": r["dob"],
            "country": r["country_of_citizenship"],
            "risk_level": risk,
            "match_data": true_matches,
            "raw_json": results_raw
        })

    conn, cur = get_db()
    for row in batch_results:
        cur.execute("""
            INSERT INTO results
                (batch_id, first_name, last_name, dob, country_of_citizenship,
                 risk_level, match_data, raw_json)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            row["batch_id"],
            row["first_name"],
            row["last_name"],
            row["dob"],
            row["country"],
            row["risk_level"],
            json.dumps(row["match_data"]),
            json.dumps(row["raw_json"]),
        ))
    conn.commit()
    cur.close()
    conn.close()


# =====================================================================
# FINISH / RESULTS
# =====================================================================

@app.route("/finish/<int:batch_id>")
@login_required
def finish(batch_id):
    rows = session.get(f"batch_{batch_id}_rows", [])
    process_batch(batch_id, rows)
    return redirect(f"/results/{batch_id}")


@app.route("/results/<int:batch_id>")
@login_required
def results(batch_id):
    return redirect("/dashboard")
